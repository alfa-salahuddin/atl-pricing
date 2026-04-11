"""
Excel template upload helpers.
Validates headers, parses rows, and returns clean data or error lists.
"""
import pandas as pd
from io import BytesIO


TEMPLATE_HEADERS = {
    "supplier_price_update": [
        "item_code", "supplier_code", "cost_currency",
        "cost_price", "discount_pct", "cost_additions",
        "effective_date", "notes",
    ],
    "new_products": [
        "item_code", "product_category", "hs_code", "product_name",
        "packing", "uom", "origin", "supplier_code",
        "cost_currency", "cost_price", "discount_pct",
        "cost_additions", "ctn_cbm", "ctn_weight", "margin_pct",
    ],
    "customers": [
        "cust_code", "name", "address", "email",
        "contact_person", "phone", "country",
    ],
}

# Only these fields are truly required — all others are optional
REQUIRED_FIELDS = {
    "supplier_price_update": [
        "item_code", "supplier_code", "cost_currency",
        "cost_price", "effective_date",
    ],
    "new_products": [
        "item_code", "product_category", "product_name",
        "packing", "uom", "origin", "supplier_code",
        "cost_currency", "cost_price", "margin_pct",
    ],
    "customers": [
        "cust_code", "name", "country",
    ],
}

NUMERIC_FIELDS = [
    "cost_price", "discount_pct", "cost_additions",
    "ctn_cbm", "ctn_weight", "margin_pct",
]


def _is_blank(val) -> bool:
    """Return True if value is blank, NaN, or the string 'nan'."""
    if val is None:
        return True
    if pd.isna(val):
        return True
    if str(val).strip().lower() in ("", "nan", "none", "n/a", "na"):
        return True
    return False


def validate_and_parse(
    file_bytes: bytes,
    template_type: str,
) -> tuple[list[dict], list[str]]:
    """
    Reads an uploaded Excel file and validates it.
    Returns (rows, errors).
    - rows:   list of clean dicts, one per valid data row
    - errors: list of human-readable error strings for invalid rows
              (valid rows are still returned even if some rows have errors)
    """
    errors = []
    rows   = []

    try:
        df = pd.read_excel(BytesIO(file_bytes), dtype=str)
    except Exception as e:
        return [], [f"Could not read file: {e}"]

    if df.empty:
        return [], ["The file appears to be empty — no data rows found."]

    # Normalise column names: lowercase, strip spaces, replace spaces with _
    df.columns = [
        str(c).strip().lower().replace(" ", "_")
        for c in df.columns
    ]

    # Drop completely empty rows
    df = df.dropna(how="all")

    expected     = TEMPLATE_HEADERS.get(template_type, [])
    required     = REQUIRED_FIELDS.get(template_type, [])

    # Check all expected columns exist — missing columns are fatal
    missing_cols = [h for h in expected if h not in df.columns]
    if missing_cols:
        return [], [
            f"Missing columns: {', '.join(missing_cols)}. "
            f"Please download the latest template and use it."
        ]

    for idx, row in df.iterrows():
        row_num    = idx + 2   # Excel row (1-indexed header row + data offset)
        row_errors = []

        # Check required fields
        for field in required:
            val = row.get(field, "")
            if _is_blank(val):
                row_errors.append(f"Row {row_num}: '{field}' is required but is blank")

        if row_errors:
            errors.extend(row_errors)
            # Still try to parse the row so we can show it in the debug output
            # but mark it as having errors — bulk upload will skip it

        # Build clean dict for ALL columns (required + optional)
        clean = {}
        for col in expected:
            val = row.get(col, "")
            if _is_blank(val):
                clean[col] = None
            else:
                clean[col] = str(val).strip()

        # Cast numeric fields to float where possible
        for f in NUMERIC_FIELDS:
            if f in clean and clean[f] is not None:
                try:
                    val = float(clean[f])
                    # Excel stores percentage-formatted cells as decimals
                    # e.g. 20% becomes 0.2 — detect and convert back to 20.0
                    # Applies to pct fields only: if value < 1 and field is a % field
                    PCT_FIELDS = ["discount_pct", "margin_pct"]
                    if f in PCT_FIELDS and 0 < val < 1:
                        val = round(val * 100, 6)
                    clean[f] = val
                except ValueError:
                    errors.append(
                        f"Row {row_num}: '{f}' must be a number, "
                        f"got '{clean[f]}'"
                    )
                    clean[f] = None

        # Only add row if it had no errors
        if not row_errors:
            rows.append(clean)

    return rows, errors


def get_template_dataframe(template_type: str) -> pd.DataFrame:
    """Returns an empty DataFrame with the correct headers for download."""
    headers = TEMPLATE_HEADERS.get(template_type, [])
    sample  = {h: "" for h in headers}
    # Add one sample row with hints
    hints = {
        "item_code":        "e.g. ATL1001",
        "product_category": "e.g. Snacks",
        "hs_code":          "e.g. 19059010 (optional)",
        "product_name":     "e.g. Twisties Chicken 60g",
        "packing":          "e.g. 24 x 60g / ctn",
        "uom":              "e.g. CTN",
        "origin":           "e.g. Malaysia",
        "supplier_code":    "e.g. MCA",
        "cost_currency":    "e.g. MYR",
        "cost_price":       "e.g. 8.50",
        "discount_pct":     "e.g. 0 (optional)",
        "cost_additions":   "e.g. 0 (optional)",
        "ctn_cbm":          "e.g. 0.042 (optional)",
        "ctn_weight":       "e.g. 14.5 (optional)",
        "margin_pct":       "e.g. 18",
        "cust_code":        "e.g. SEY001",
        "name":             "e.g. Victoria Trading Co",
        "address":          "e.g. Victoria, Mahe (optional)",
        "email":            "e.g. buyer@example.com (optional)",
        "contact_person":   "e.g. John Smith (optional)",
        "phone":            "e.g. +248 123456 (optional)",
        "country":          "e.g. Seychelles",
        "supplier_code":    "e.g. MCA",
        "effective_date":   "e.g. 2026-04-11",
        "notes":            "e.g. Q2 price revision (optional)",
    }
    row = {h: hints.get(h, "") for h in headers}
    df  = pd.DataFrame([row], columns=headers)
    return df


def dataframe_to_excel_bytes(
    df: pd.DataFrame,
    sheet_name: str = "Template",
) -> bytes:
    """Converts a DataFrame to formatted Excel bytes for download."""
    from openpyxl.styles import Font, PatternFill, Alignment
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        # Style the header row
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = PatternFill("solid", fgColor="1F5C0F")
            cell.alignment = Alignment(horizontal="center")

        # Auto-size columns
        for col in ws.columns:
            max_len = max(
                len(str(cell.value or "")) for cell in col
            )
            ws.column_dimensions[col[0].column_letter].width = max(
                max_len + 4, 16
            )

    return buf.getvalue()
