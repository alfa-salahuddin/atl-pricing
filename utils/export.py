"""
Excel export helpers.
Generates formatted .xlsx files for price lists and proforma invoices.
"""
import io
from datetime import date
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


ATL_GREEN  = "1F5C0F"
ATL_LGREY  = "F2F2F2"
ATL_WHITE  = "FFFFFF"

COMPANY = {
    "name":    "Alfa Tradelinks Pte Ltd",
    "address": "Singapore",
    "uen":     "",
    "email":   "",
    "phone":   "",
}


def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _header_style(ws, row, col, value, bold=True, bg=ATL_GREEN, fg=ATL_WHITE, size=11):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(name="Calibri", bold=bold, color=fg, size=size)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _thin_border()
    return cell


def _data_style(ws, row, col, value, number_format=None, bold=False, bg=ATL_WHITE):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(name="Calibri", size=10, bold=bold)
    cell.alignment = Alignment(vertical="center")
    cell.border    = _thin_border()
    cell.fill      = PatternFill("solid", fgColor=bg)
    if number_format:
        cell.number_format = number_format
    return cell


# ---------------------------------------------------------------------------
# Price List Excel
# ---------------------------------------------------------------------------
def export_price_list(rows: list[dict], meta: dict) -> bytes:
    """
    rows: list of dicts — one per product line
    meta: {quot_id, supplier_name, port_name, incoterm, validity_date, notes}
    Returns bytes of .xlsx file.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Price List"
    ws.sheet_view.showGridLines = False

    # --- Company header ---
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value     = COMPANY["name"]
    c.font      = Font(name="Calibri", bold=True, size=14, color=ATL_GREEN)
    c.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A2:H2")
    ws["A2"].value     = "PRICE LIST"
    ws["A2"].font      = Font(name="Calibri", bold=True, size=12)
    ws["A2"].alignment = Alignment(horizontal="left")

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18

    # --- Meta info ---
    info_row = 3
    info = [
        ("Reference",   meta.get("quot_id", "")),
        ("Supplier",    meta.get("supplier_name", "")),
        ("Port",        meta.get("port_name", "")),
        ("Incoterm",    meta.get("incoterm", "")),
        ("Valid until", meta.get("validity_date", "")),
        ("Date",        str(date.today())),
    ]
    for i, (label, value) in enumerate(info):
        col = (i % 3) * 2 + 1
        r   = info_row + (i // 3)
        ws.cell(row=r, column=col,   value=label).font = Font(bold=True, size=9, color="666666")
        ws.cell(row=r, column=col+1, value=value).font = Font(size=9)

    # --- Column headers ---
    hdr_row = info_row + 3
    headers = ["#", "Item code", "Product name", "Packing", "UOM", "Origin", "FOB price (SGD)", "CTN CBM", "CTN weight (kg)"]
    widths  = [4,    14,          35,              20,        8,     12,        16,                10,        16]

    for col_idx, (h, w) in enumerate(zip(headers, widths), start=1):
        _header_style(ws, hdr_row, col_idx, h)
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.row_dimensions[hdr_row].height = 30

    # --- Data rows ---
    for i, row in enumerate(rows, start=1):
        dr  = hdr_row + i
        bg  = ATL_WHITE if i % 2 == 0 else ATL_LGREY
        _data_style(ws, dr, 1, i,                           bg=bg)
        _data_style(ws, dr, 2, row.get("item_code", ""),    bg=bg)
        _data_style(ws, dr, 3, row.get("product_name", ""), bg=bg)
        _data_style(ws, dr, 4, row.get("packing", ""),      bg=bg)
        _data_style(ws, dr, 5, row.get("uom", ""),          bg=bg)
        _data_style(ws, dr, 6, row.get("origin", ""),       bg=bg)
        _data_style(ws, dr, 7, row.get("fob_price_sgd", 0), number_format='"SGD "#,##0.00', bold=True, bg=bg)
        _data_style(ws, dr, 8, row.get("ctn_cbm", ""),      bg=bg)
        _data_style(ws, dr, 9, row.get("ctn_weight", ""),   bg=bg)
        ws.row_dimensions[dr].height = 16

    # --- Notes footer ---
    if meta.get("notes"):
        note_row = hdr_row + len(rows) + 2
        ws.merge_cells(f"A{note_row}:I{note_row}")
        ws[f"A{note_row}"].value = f"Notes: {meta['notes']}"
        ws[f"A{note_row}"].font  = Font(size=9, italic=True, color="666666")

    # --- Freeze panes ---
    ws.freeze_panes = f"A{hdr_row + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Proforma Invoice Excel
# ---------------------------------------------------------------------------
def export_proforma_invoice(rows: list[dict], meta: dict) -> bytes:
    """
    rows: list of dicts — one per line item (includes qty_ctns)
    meta: {quot_id, cust_name, cust_address, supplier_name, port_name,
           incoterm, validity_date, shipping_line, notes}
    Returns bytes of .xlsx file.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proforma Invoice"
    ws.sheet_view.showGridLines = False

    # Company header
    ws.merge_cells("A1:I1")
    ws["A1"].value     = COMPANY["name"]
    ws["A1"].font      = Font(name="Calibri", bold=True, size=14, color=ATL_GREEN)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A2:I2")
    ws["A2"].value     = "PROFORMA INVOICE"
    ws["A2"].font      = Font(name="Calibri", bold=True, size=12)
    ws["A2"].alignment = Alignment(horizontal="left")

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18

    # Meta
    info_row = 3
    info = [
        ("PI Reference", meta.get("quot_id", "")),
        ("Customer",     meta.get("cust_name", "")),
        ("Address",      meta.get("cust_address", "")),
        ("Supplier",     meta.get("supplier_name", "")),
        ("Port",         meta.get("port_name", "")),
        ("Incoterm",     meta.get("incoterm", "")),
        ("Shipping line",meta.get("shipping_line", "")),
        ("Valid until",  meta.get("validity_date", "")),
        ("Date",         str(date.today())),
    ]
    for i, (label, value) in enumerate(info):
        col = (i % 3) * 2 + 1
        r   = info_row + (i // 3)
        ws.cell(row=r, column=col,   value=label).font = Font(bold=True, size=9, color="666666")
        ws.cell(row=r, column=col+1, value=value).font = Font(size=9)

    # Column headers
    hdr_row = info_row + 5
    headers = ["#", "Item code", "Product name", "Packing", "UOM", "Qty (ctns)", "Unit price (SGD)", "Amount (SGD)", "CTN CBM", "Total CBM"]
    widths  = [4,    14,          32,              18,        7,     10,            17,                 14,             10,         10]

    for col_idx, (h, w) in enumerate(zip(headers, widths), start=1):
        _header_style(ws, hdr_row, col_idx, h)
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.row_dimensions[hdr_row].height = 30

    # Data rows
    total_amount = 0.0
    total_cbm    = 0.0
    for i, row in enumerate(rows, start=1):
        dr          = hdr_row + i
        bg          = ATL_WHITE if i % 2 == 0 else ATL_LGREY
        qty         = row.get("qty_ctns") or 0
        unit_price  = row.get("fob_price_sgd", 0)
        amount      = round(qty * unit_price, 2)
        cbm_per_ctn = row.get("ctn_cbm") or 0
        total_line_cbm = round(qty * cbm_per_ctn, 4)

        total_amount += amount
        total_cbm    += total_line_cbm

        _data_style(ws, dr, 1,  i,                           bg=bg)
        _data_style(ws, dr, 2,  row.get("item_code", ""),    bg=bg)
        _data_style(ws, dr, 3,  row.get("product_name", ""), bg=bg)
        _data_style(ws, dr, 4,  row.get("packing", ""),      bg=bg)
        _data_style(ws, dr, 5,  row.get("uom", ""),          bg=bg)
        _data_style(ws, dr, 6,  qty,                         number_format="#,##0",        bg=bg)
        _data_style(ws, dr, 7,  unit_price,                  number_format='"SGD "#,##0.00', bg=bg)
        _data_style(ws, dr, 8,  amount,                      number_format='"SGD "#,##0.00', bold=True, bg=bg)
        _data_style(ws, dr, 9,  cbm_per_ctn,                 number_format="#,##0.0000",   bg=bg)
        _data_style(ws, dr, 10, total_line_cbm,              number_format="#,##0.0000",   bg=bg)
        ws.row_dimensions[dr].height = 16

    # Totals row
    tot_row = hdr_row + len(rows) + 1
    ws.merge_cells(f"A{tot_row}:G{tot_row}")
    c = ws[f"A{tot_row}"]
    c.value     = "TOTAL"
    c.font      = Font(bold=True, size=11)
    c.alignment = Alignment(horizontal="right")
    c.fill      = PatternFill("solid", fgColor=ATL_LGREY)

    _data_style(ws, tot_row, 8,  round(total_amount, 2), number_format='"SGD "#,##0.00', bold=True, bg=ATL_LGREY)
    _data_style(ws, tot_row, 10, round(total_cbm, 4),    number_format="#,##0.0000",    bold=True, bg=ATL_LGREY)

    # Notes
    if meta.get("notes"):
        note_row = tot_row + 2
        ws.merge_cells(f"A{note_row}:J{note_row}")
        ws[f"A{note_row}"].value = f"Notes: {meta['notes']}"
        ws[f"A{note_row}"].font  = Font(size=9, italic=True, color="666666")

    ws.freeze_panes = f"A{hdr_row + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

